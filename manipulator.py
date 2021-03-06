from openpyxl import load_workbook, worksheet, Workbook
from collections import namedtuple

import os
import sys, getopt

def main(argv):
    inputfile = ''
    outputfile = ''
    try:
        opts, args = getopt.getopt(argv,"hi:o:",["ifile=","ofile="])
    except getopt.GetoptError:
        print ('manipulator.py -i <inputfile> -o <outputfile>')
        sys.exit(2)
    for opt, arg in opts:
        if opt == '-h':
            print ('manipulator.py -i <inputfile> -o <outputfile>')
            sys.exit()
        elif opt in ("-i", "--ifile"):
            inputfile = arg
        elif opt in ("-o", "--ofile"):
            outputfile = arg

    theFile = load_workbook(inputfile)
    allSheetNames = theFile.sheetnames

    excel_file = Workbook()
    excel_sheet = excel_file.create_sheet(title='Result', index=0)

    Influencer = namedtuple("Influencer", "name contact occupation party")
    VillageDetail = namedtuple("VillageDetails", "uid district ac block village")
    HotspotDetails = namedtuple("HotspotDetails", "sc people village village_people")

    for sheet in allSheetNames:
        seen = set()
        result = {}
        villageDetails = {}
        hotspot = {}
        
        currentSheet = theFile[sheet]
        
        for row in range(2, currentSheet.max_row + 1):
            uidVillage = currentSheet.cell(row = row, column = 3).value
            if uidVillage not in seen :
                seen.add(uidVillage)
                result[uidVillage] = []
                villageDetails[uidVillage] = VillageDetail(
                    currentSheet.cell(row = row, column = 3).value, 
                    currentSheet.cell(row = row, column = 4).value, 
                    currentSheet.cell(row = row, column = 5).value, 
                    currentSheet.cell(row = row, column = 6).value,
                    currentSheet.cell(row = row, column = 7).value)

            # Add influencers
            Influencers = []
            Influencers.append(Influencer(currentSheet.cell(row = row, column = 8).value, currentSheet.cell(row = row, column = 9).value, currentSheet.cell(row = row, column = 10).value, currentSheet.cell(row = row, column = 11).value))
            Influencers.append(Influencer(currentSheet.cell(row = row, column = 12).value, currentSheet.cell(row = row, column = 13).value, currentSheet.cell(row = row, column = 14).value, currentSheet.cell(row = row, column = 15).value))
            Influencers.append(Influencer(currentSheet.cell(row = row, column = 16).value, currentSheet.cell(row = row, column = 17).value, currentSheet.cell(row = row, column = 18).value, currentSheet.cell(row = row, column = 19).value))
            Influencers.append(Influencer(currentSheet.cell(row = row, column = 20).value, currentSheet.cell(row = row, column = 21).value, currentSheet.cell(row = row, column = 22).value, currentSheet.cell(row = row, column = 23).value))
            Influencers.append(Influencer(currentSheet.cell(row = row, column = 24).value, currentSheet.cell(row = row, column = 25).value, currentSheet.cell(row = row, column = 26).value, currentSheet.cell(row = row, column = 27).value))
            Influencers.append(Influencer(currentSheet.cell(row = row, column = 28).value, currentSheet.cell(row = row, column = 29).value, currentSheet.cell(row = row, column = 30).value, currentSheet.cell(row = row, column = 31).value))
            Influencers.append(Influencer(currentSheet.cell(row = row, column = 32).value, currentSheet.cell(row = row, column = 33).value, currentSheet.cell(row = row, column = 34).value, currentSheet.cell(row = row, column = 35).value))
            Influencers.append(Influencer(currentSheet.cell(row = row, column = 36).value, currentSheet.cell(row = row, column = 37).value, currentSheet.cell(row = row, column = 38).value, currentSheet.cell(row = row, column = 39).value))
            Influencers.append(Influencer(currentSheet.cell(row = row, column = 40).value, currentSheet.cell(row = row, column = 41).value, currentSheet.cell(row = row, column = 42).value, currentSheet.cell(row = row, column = 43).value))
            Influencers.append(Influencer(currentSheet.cell(row = row, column = 44).value, currentSheet.cell(row = row, column = 45).value, currentSheet.cell(row = row, column = 46).value, currentSheet.cell(row = row, column = 47).value))

            Influencers = [x for x in Influencers if x.name]
            
            result[uidVillage] += Influencers
            
            # Hotspot detail aggregation :
            existing_details = hotspot.get(uidVillage)
            if existing_details is None :
                existing_details = HotspotDetails('', '', '', '')

            try :
                sc = str(currentSheet.cell(row = row, column = 48).value or '')
            except :
                print("Exception caught reading hotspot sc for village uid " + str(uidVillage) + ", where sc was " + currentSheet.cell(row = row, column = 48).value)

            people = (currentSheet.cell(row = row, column = 49).value or '')
            village = str(currentSheet.cell(row = row, column = 50).value or '')
            village_people = str(currentSheet.cell(row = row, column = 51).value or '')

            new_sc = existing_details.sc
            new_people = existing_details.people
            new_village = existing_details.village
            new_village_people = existing_details.village_people

            #Only add sc and people if sc is not null. Similarly only add village and village_people if village is not null
            if (sc != '') :
                new_sc = new_sc + sc.strip() + ","
                try:
                    new_people = new_people + str(people).strip() + ","
                except:
                    print("Exception caught reading hotspot sc people for village uid " + str(uidVillage) + ", where people was " + people)
            
            if (village != '') :
                new_village = new_village + village.strip() + ","
                new_village_people = new_village_people + str(village_people).strip() + ","
        
            hotspot[uidVillage] = HotspotDetails(new_sc, new_people, new_village, new_village_people)
            
    # Now lets start writing the new output excel sheet data
    header = ["UID", "District", "AC", "Block", "Village", "Hotspot SC", "Hotspot_SC_People", "Hotspot_Village", "Hotspot_Village_People", "Influencer Name", "Influencer Contact", "Influencer Occupation", "Influencer Party"]
    excel_sheet.append(header)

    for village in seen :
        rowData = []
        details = villageDetails[village]
        influencers = result[village]
        hotspotData = hotspot[village]

        #Start with village data
        rowData.append(details.uid)
        rowData.append(details.district)
        rowData.append(details.ac)
        rowData.append(details.block)
        rowData.append(details.village)
        new_sc = hotspotData.sc
        if (new_sc != '') :
            new_sc = new_sc[:-1]
    
        new_people = hotspotData.people
        if (new_people != '') :
            new_people = new_people[:-1]

        new_village = hotspotData.village
        if (new_village != '') :
            new_village = new_village[:-1]

        new_village_people = hotspotData.village_people
        if (new_village_people != '') :
            new_village_people = new_village_people[:-1]

        # Add hotspots now
        rowData.append(new_sc)
        rowData.append(new_people)
        rowData.append(new_village)
        rowData.append(new_village_people)

        # Add influencers now
        for influencer in influencers :
            rowData.append(influencer.name)
            rowData.append(influencer.contact)
            rowData.append(influencer.occupation)
            rowData.append(influencer.party)
    
        excel_sheet.append(rowData)

    excel_file.save(outputfile)
    print("A new file created for village count " + str(len(villageDetails)))

if __name__ == "__main__":
   main(sys.argv[1:])

    


